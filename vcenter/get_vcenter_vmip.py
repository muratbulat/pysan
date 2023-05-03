from com.vmware.vcenter.vm_client import Power
from com.vmware.vcenter.vm_client import Network

# Connect to vCenter server
from vmware.vapi.vsphere.client import create_vsphere_client
client = create_vsphere_client(server="<VCENTER_SERVER_IP>", username="<USERNAME>", password="<PASSWORD>")

# Get the power and network information of all virtual machines
power = client.vcenter.vm.Power
network = client.vcenter.vm.Network
vms = client.vcenter.VM.list()

# Create an Excel workbook and worksheet
from openpyxl import Workbook
workbook = Workbook()
worksheet = workbook.active
worksheet.title = "VM IP Addresses"

# Write the headers
worksheet.cell(row=1, column=1, value="VM Name")
worksheet.cell(row=1, column=2, value="IP Address")

# Iterate through all virtual machines and write their IP addresses to the worksheet
for i, vm in enumerate(vms):
    vm_name = vm.name
    ip_address = None
    for nic_info in network.list(vm.vm):
        if nic_info.nic_type == "VMXNET3":
            nic = nic_info.nic
            if nic.backing.network_type == "STANDARD_PORTGROUP":
                ip_address = nic.ip_address.ip_address
    worksheet.cell(row=i+2, column=1, value=vm_name)
    worksheet.cell(row=i+2, column=2, value=ip_address)

# Save the Excel file
workbook.save(filename="vm_ip_addresses.xlsx")
